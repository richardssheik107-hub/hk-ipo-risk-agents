"""Audit governed Market-X Extended sources across the official 438 cohort.

The detailed 438-row artifact is written below ignored local data.  Only the
aggregate summary is intended for version control.  No 2025 outcome is read.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from ipo_risk.market.csmar_hsi import CSMAR_HSI_REFERENCE_ID, CSMARHSIProvider
from ipo_risk.market.features import PreListingMarketFeatureEngine
from ipo_risk.market.labels import MarketLabelGenerator
from ipo_risk.market.official_market_sources import (
    ExternalMarketSourceManifest,
    OfficialHSCIProvider,
    OfficialHKEXTurnoverProvider,
)
from ipo_risk.providers.competition_market import CompetitionCSVMarketDataProvider
from ipo_risk.schemas.market import (
    IPOMarketMetadata,
    MarketDataProvenance,
    MarketLabelHorizon,
    MarketOutcomeLabel,
    expected_market_split,
)
from ipo_risk.schemas.market_features import (
    MARKET_RAW_FEATURE_ORDER,
    MarketFeatureAvailability,
    PreListingMarketFeatureContext,
    PriorIPOReference,
)


EXPECTED_CASES = 438
MAPPING_PIT_BLOCKED = "INDUSTRY_MAPPING_PIT_BLOCKED"
HISTORY_REASONS = {
    "BENCHMARK_HISTORY_NOT_YET_STARTED",
    "INSUFFICIENT_5D_HISTORY",
    "INSUFFICIENT_20D_HISTORY",
}


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _canonical_hash(value: Any) -> str:
    payload = json.dumps(
        value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
    ).encode("utf-8")
    return hashlib.sha256(payload).hexdigest()


def _date_text(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "").strip()[:10]


def _symbol(value: object) -> str:
    text = str(value or "").strip()
    return text.zfill(5) if text else ""


def load_mapping(path: Path) -> dict[str, dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    mapping = {row["industry_code"].zfill(2): row for row in rows}
    if len(rows) != 12 or len(mapping) != 12:
        raise ValueError("HSICS mapping draft must contain 12 unique codes")
    if any(
        row["effective_from"]
        or row["effective_to"]
        or row["pit_status"] != "PIT_BLOCKED"
        for row in rows
    ):
        raise ValueError("HSICS mapping draft must preserve null dates and PIT block")
    return mapping


def load_static_industry_fields(
    workbook_path: Path,
    case_ids: set[str],
) -> tuple[dict[str, dict[str, str]], dict[str, Any]]:
    """Read the delivered static workbook only for a temporal-semantics audit."""

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    required = {"Merged_Official_Data", "IPO_565_Match", "Field_Dictionary"}
    if not required.issubset(workbook.sheetnames):
        raise ValueError("official workbook is missing required audit sheets")

    dictionary_rows = list(workbook["Field_Dictionary"].values)
    dictionary_header = {name: index for index, name in enumerate(dictionary_rows[0])}
    source_by_field = {
        str(row[dictionary_header["OfficialFieldName"]]): str(
            row[dictionary_header["SourceTable(s)"]] or ""
        )
        for row in dictionary_rows[1:]
    }
    industry_fields = {
        "IndustryCode",
        "INDUSTRYNAME",
        "IndustryCode2",
        "IndustryName2",
    }
    if any(source_by_field.get(field) != "Institution" for field in industry_fields):
        raise ValueError("industry fields are not consistently Institution-source fields")

    merged_rows = list(workbook["Merged_Official_Data"].values)
    merged_header = {name: index for index, name in enumerate(merged_rows[0])}
    merged_by_key = {
        (
            _symbol(row[merged_header["Symbol"]]),
            _date_text(row[merged_header["ListedDate"]]),
            str(row[merged_header["InstitutionID"]] or "").strip(),
        ): row
        for row in merged_rows[3:]
        if row[merged_header["Symbol"]] is not None
    }

    match_rows = list(workbook["IPO_565_Match"].values)
    match_header = {name: index for index, name in enumerate(match_rows[0])}
    output: dict[str, dict[str, str]] = {}
    declare_after_listing = 0
    declare_before_listing = 0
    declare_unparsed = 0
    code2_prefix_mismatch = 0
    for match in match_rows[1:]:
        case_id = str(match[match_header["CaseID"]] or "").strip()
        if case_id not in case_ids:
            continue
        if str(match[match_header["MatchStatus"]] or "") != "matched":
            raise ValueError(f"official case is not matched in workbook: {case_id}")
        key = (
            _symbol(match[match_header["MatchedSymbol"]]),
            _date_text(match[match_header["MatchedListedDate"]]),
            str(match[match_header["MatchedInstitutionID"]] or "").strip(),
        )
        try:
            row = merged_by_key[key]
        except KeyError as exc:
            raise ValueError(f"official workbook join failed for {case_id}") from exc
        code_raw = row[merged_header["IndustryCode"]]
        code = str(code_raw).strip().zfill(2) if code_raw is not None else ""
        code2 = str(row[merged_header["IndustryCode2"]] or "").strip()
        if code and code2 and not code2.startswith(code):
            code2_prefix_mismatch += 1
        listing_date = date.fromisoformat(key[1])
        try:
            declare_date = date.fromisoformat(
                _date_text(row[merged_header["DeclareDate"]])
            )
            declare_after_listing += int(declare_date > listing_date)
            declare_before_listing += int(declare_date < listing_date)
        except ValueError:
            declare_unparsed += 1
        output[case_id] = {
            "industry_code": code,
            "industry_name": str(row[merged_header["INDUSTRYNAME"]] or "").strip(),
            "industry_code2": code2,
            "industry_name2": str(row[merged_header["IndustryName2"]] or "").strip(),
            "sector_ind_name": str(row[merged_header["SectorIndName"]] or "").strip(),
            "declare_date": _date_text(row[merged_header["DeclareDate"]]),
        }
    if set(output) != case_ids:
        raise ValueError(
            f"industry workbook cohort join drift: {len(output)} != {len(case_ids)}"
        )
    missing_classification = sum(not item["industry_code"] for item in output.values())
    return output, {
        "workbook_sha256": _sha256(workbook_path),
        "workbook_created": (
            workbook.properties.created.isoformat()
            if workbook.properties.created is not None
            else None
        ),
        "field_source_semantics": "Institution static merged record",
        "industry_field_sources": {
            field: source_by_field[field] for field in sorted(industry_fields)
        },
        "declare_date_after_listing_count": declare_after_listing,
        "declare_date_before_listing_count": declare_before_listing,
        "declare_date_unparsed_count": declare_unparsed,
        "industry_code2_prefix_mismatch_count": code2_prefix_mismatch,
        "missing_classification_count": missing_classification,
        "temporal_semantics": "NO_CLASSIFICATION_EFFECTIVE_DATE_OR_LISTING_TIME_SNAPSHOT_ASSERTION",
        "pit_safe": False,
        "production_status": MAPPING_PIT_BLOCKED,
    }


def build_labels_and_prior_ipos(
    metadata: tuple[IPOMarketMetadata, ...],
    provider: CompetitionCSVMarketDataProvider,
) -> tuple[tuple[PriorIPOReference, ...], tuple[MarketOutcomeLabel, ...]]:
    generator = MarketLabelGenerator()
    prior_ipos: list[PriorIPOReference] = []
    labels: list[MarketOutcomeLabel] = []
    for item in metadata:
        if item.cohort_year >= 2025 or item.listing_date is None:
            raise ValueError("2025 Blind or missing listing date entered Extended audit")
        prior_ipos.append(
            PriorIPOReference(
                case_id=item.case_id,
                stock_code=item.stock_code,
                cohort_year=item.cohort_year,
                listing_date=item.listing_date,
                dataset_split=expected_market_split(item.cohort_year),
                official_ipo_universe_member=item.official_ipo_universe_member,
                security_type=item.security_type,
                modeling_eligibility=item.modeling_eligibility,
                eligibility_reason=item.eligibility_reason,
                eligibility_policy_version=item.eligibility_policy_version,
                provenance=item.provenance,
            )
        )
        generated = generator.generate(
            item, provider.get_daily_bars(item.stock_code)
        )
        labels.extend(
            label
            for label in generated
            if label.horizon in {
                MarketLabelHorizon.ONE_DAY,
                MarketLabelHorizon.FIVE_DAYS,
            }
        )
    return tuple(prior_ipos), tuple(labels)


def _context(item: IPOMarketMetadata, hsi_provider: CSMARHSIProvider) -> PreListingMarketFeatureContext:
    assert item.listing_date is not None
    return PreListingMarketFeatureContext(
        case_id=item.case_id,
        stock_code=item.stock_code,
        cohort_year=item.cohort_year,
        listing_date=item.listing_date,
        dataset_split=expected_market_split(item.cohort_year),
        benchmark_reference_id=CSMAR_HSI_REFERENCE_ID,
        industry_reference_id=None,
        source="governed official Market-X Extended sources",
        provenance=MarketDataProvenance(
            source="governed_market_x_extended",
            dataset_version="v04_c_industry_turnover_integration_v1",
            source_record_id=item.case_id,
            metadata={
                "hsi_source_version": hsi_provider.manifest.source_version(),
                "industry_mapping_pit_status": MAPPING_PIT_BLOCKED,
            },
        ),
    )


def _conditional_reason(
    *, industry_code: str, benchmark_id: str, observation_count: int, sessions: int
) -> str:
    if not industry_code:
        return "MISSING_INDUSTRY_CLASSIFICATION"
    if not benchmark_id:
        return "MISSING_INDUSTRY_MAPPING"
    if observation_count == 0:
        return "BENCHMARK_HISTORY_NOT_YET_STARTED"
    if observation_count < sessions + 1:
        return f"INSUFFICIENT_{sessions}D_HISTORY"
    return ""


def materialize(
    metadata: tuple[IPOMarketMetadata, ...],
    industry_fields: dict[str, dict[str, str]],
    mapping: dict[str, dict[str, str]],
    hsi_provider: CSMARHSIProvider,
    hsci_provider: OfficialHSCIProvider,
    turnover_provider: OfficialHKEXTurnoverProvider,
    prior_ipos: tuple[PriorIPOReference, ...],
    labels: tuple[MarketOutcomeLabel, ...],
    *,
    provider_cutoffs: bool,
) -> list[dict[str, Any]]:
    engine = PreListingMarketFeatureEngine()
    all_hsi = tuple(hsi_provider.iter_all_bars())
    all_turnover = tuple(turnover_provider.iter_all_observations())
    output: list[dict[str, Any]] = []
    for item in metadata:
        assert item.listing_date is not None
        context = _context(item, hsi_provider)
        benchmark_bars = (
            hsi_provider.get_benchmark_bars(
                CSMAR_HSI_REFERENCE_ID, end_date_exclusive=item.listing_date
            )
            if provider_cutoffs
            else all_hsi
        )
        activity = (
            turnover_provider.get_market_activity(end_date_exclusive=item.listing_date)
            if provider_cutoffs
            else all_turnover
        )
        snapshot = engine.build(
            context,
            benchmark_bars=benchmark_bars,
            industry_bars=None,
            activity_observations=activity,
            prior_ipos=prior_ipos,
            prior_outcomes=labels,
        )
        features = {feature.name: feature for feature in snapshot.features}
        static = industry_fields[item.case_id]
        provisional_mapping = mapping.get(static["industry_code"])
        benchmark_id = (
            provisional_mapping["benchmark_id"] if provisional_mapping else ""
        )
        hsci_bars = (
            hsci_provider.get_industry_bars(
                benchmark_id, end_date_exclusive=item.listing_date
            )
            if benchmark_id
            else ()
        )
        # Align the conditional research calculation with the HSI observation date,
        # exactly as the frozen engine would.  It is not a production feature.
        through = snapshot.observation_date
        aligned_hsci = tuple(
            bar for bar in hsci_bars if through is None or bar.trading_date <= through
        )
        conditional_values: dict[str, str] = {}
        for sessions in (5, 20):
            reason = _conditional_reason(
                industry_code=static["industry_code"],
                benchmark_id=benchmark_id,
                observation_count=len(aligned_hsci),
                sessions=sessions,
            )
            conditional_values[f"conditional_industry_{sessions}d_available"] = not reason
            conditional_values[f"conditional_industry_{sessions}d_missing_reason"] = reason
            conditional_values[f"conditional_industry_return_{sessions}d"] = (
                str(aligned_hsci[-1].close / aligned_hsci[-(sessions + 1)].close - 1)
                if not reason
                else ""
            )
        row: dict[str, Any] = {
            "case_id": item.case_id,
            "stock_code": item.stock_code,
            "listing_date": item.listing_date.isoformat(),
            "listing_year": item.cohort_year,
            "dataset_split": expected_market_split(item.cohort_year).value,
            "industry_code": static["industry_code"],
            "industry_name": static["industry_name"],
            "provisional_benchmark_id": benchmark_id,
            "static_mapping_available": bool(benchmark_id),
            "mapping_available": False,
            "industry_return_5d": "",
            "industry_5d_available": False,
            "industry_5d_missing_reason": (
                "MISSING_INDUSTRY_CLASSIFICATION"
                if not static["industry_code"]
                else MAPPING_PIT_BLOCKED
            ),
            "industry_return_20d": "",
            "industry_20d_available": False,
            "industry_20d_missing_reason": (
                "MISSING_INDUSTRY_CLASSIFICATION"
                if not static["industry_code"]
                else MAPPING_PIT_BLOCKED
            ),
            "benchmark_observation_date": (
                aligned_hsci[-1].trading_date.isoformat() if aligned_hsci else ""
            ),
            "benchmark_source_version": (
                aligned_hsci[-1].provenance.dataset_version if aligned_hsci else ""
            ),
            **conditional_values,
        }
        available_raw = 0
        for name in MARKET_RAW_FEATURE_ORDER:
            feature = features[name]
            available = feature.availability is MarketFeatureAvailability.AVAILABLE
            row[f"{name}__available"] = available
            row[f"{name}__missing"] = not available
            row[f"{name}__missing_reason"] = (
                feature.missing_reason.value if feature.missing_reason else ""
            )
            row[name] = str(feature.value) if available else ""
            available_raw += int(available)
        # The frozen engine reports missing_industry_mapping.  The readiness audit
        # records the stricter governance cause without changing that public enum.
        for name in ("industry_return_5d", "industry_return_20d"):
            row[f"{name}__missing_reason"] = row[
                "industry_5d_missing_reason"
                if name.endswith("5d")
                else "industry_20d_missing_reason"
            ]
        row["available_raw_feature_count"] = available_raw
        row["missing_raw_feature_count"] = len(MARKET_RAW_FEATURE_ORDER) - available_raw
        row["FULL_10_RAW_AVAILABLE"] = available_raw == 10
        row["PARTIAL_AVAILABLE"] = 0 < available_raw < 10
        output.append(row)
    output.sort(key=lambda row: row["case_id"])
    if len(output) != EXPECTED_CASES or len({row["case_id"] for row in output}) != EXPECTED_CASES:
        raise ValueError("Extended audit did not preserve exactly 438 unique cases")
    return output


def summarize(
    rows: list[dict[str, Any]],
    manifest: ExternalMarketSourceManifest,
    temporal_audit: dict[str, Any],
) -> dict[str, Any]:
    by_year: dict[int, dict[str, int]] = {}
    for year in range(2020, 2025):
        group = [row for row in rows if row["listing_year"] == year]
        by_year[year] = {
            "cases": len(group),
            "static_mapping_available": sum(row["static_mapping_available"] for row in group),
            "production_mapping_available": 0,
            "conditional_5d_available": sum(
                row["conditional_industry_5d_available"] for row in group
            ),
            "conditional_20d_available": sum(
                row["conditional_industry_20d_available"] for row in group
            ),
            "missing_because_history_starts_2021_08_19_5d": sum(
                row["conditional_industry_5d_missing_reason"] in HISTORY_REASONS
                for row in group
            ),
            "missing_because_history_starts_2021_08_19_20d": sum(
                row["conditional_industry_20d_missing_reason"] in HISTORY_REASONS
                for row in group
            ),
            "missing_because_classification_missing": sum(
                not row["industry_code"] for row in group
            ),
        }
    raw_availability = {
        name: sum(row[f"{name}__available"] for row in rows)
        for name in MARKET_RAW_FEATURE_ORDER
    }
    missing_reasons: Counter[str] = Counter()
    for row in rows:
        for name in MARKET_RAW_FEATURE_ORDER:
            reason = str(row[f"{name}__missing_reason"])
            if reason:
                missing_reasons[f"{name}:{reason}"] += 1
    supported_industry_reasons = (
        "MISSING_INDUSTRY_CLASSIFICATION",
        "MISSING_INDUSTRY_MAPPING",
        "BENCHMARK_HISTORY_NOT_YET_STARTED",
        "INSUFFICIENT_5D_HISTORY",
        "INSUFFICIENT_20D_HISTORY",
        "BENCHMARK_SOURCE_ERROR",
    )
    conditional_5d_reasons = Counter(
        row["conditional_industry_5d_missing_reason"]
        for row in rows
        if row["conditional_industry_5d_missing_reason"]
    )
    conditional_20d_reasons = Counter(
        row["conditional_industry_20d_missing_reason"]
        for row in rows
        if row["conditional_industry_20d_missing_reason"]
    )
    affected = {
        row["case_id"]
        for row in rows
        if row["conditional_industry_5d_missing_reason"] in HISTORY_REASONS
        or row["conditional_industry_20d_missing_reason"] in HISTORY_REASONS
    }
    affected_years = sorted(
        {row["listing_year"] for row in rows if row["case_id"] in affected}
    )
    development_affected = sum(
        row["case_id"] in affected and row["dataset_split"] == "development"
        for row in rows
    )
    validation_affected = sum(
        row["case_id"] in affected and row["dataset_split"] == "validation"
        for row in rows
    )
    return {
        "audit_version": "v04_c_extended_readiness_v1",
        "generated_by": "scripts/run_v04_c_extended_readiness.py",
        "official_cases": len(rows),
        "industry_taxonomy_status": "ACCEPT",
        "industry_mapping_status": "EVIDENCE_BACKED_DRAFT",
        "industry_mapping_pit_status": MAPPING_PIT_BLOCKED,
        "industry_mapping_pit_blocked": True,
        "mapping_acceptance": {
            "FIELD_SOURCE_SEMANTICS": temporal_audit["field_source_semantics"],
            "TAXONOMY_MATCH": "PASS",
            "TOP_LEVEL_CODE_MATCH": "PASS",
            "NAME_MATCH": "PASS",
            "TEMPORAL_SEMANTICS": temporal_audit["temporal_semantics"],
            "PIT_SAFE": False,
        },
        "temporal_audit": temporal_audit,
        "hsci_source_status": manifest.hsci_industry_daily_close.status,
        "hsci_series_accepted": manifest.hsci_industry_daily_close.accepted_series_count,
        "hsci_production_5d_available": 0,
        "hsci_production_20d_available": 0,
        "conditional_static_mapping_5d_available": sum(
            row["conditional_industry_5d_available"] for row in rows
        ),
        "conditional_static_mapping_20d_available": sum(
            row["conditional_industry_20d_available"] for row in rows
        ),
        "industry_coverage_by_listing_year": by_year,
        "turnover_source_status": manifest.hkex_total_market_daily_turnover.status,
        "turnover_20d_available": raw_availability["market_turnover_20d_mean"],
        "turnover_earliest_observation_date": (
            manifest.hkex_total_market_daily_turnover.coverage_start.isoformat()
        ),
        "recent_ipo_features": {
            name: raw_availability[name]
            for name in (
                "recent_ipo_break_rate",
                "recent_ipo_return_5d",
                "recent_ipo_1d_sample_count",
                "recent_ipo_5d_sample_count",
            )
        },
        "raw_feature_availability": raw_availability,
        "extended_raw_feature_count": len(MARKET_RAW_FEATURE_ORDER),
        "extended_position_count": len(MARKET_RAW_FEATURE_ORDER) * 2,
        "full_10_raw_available": sum(row["FULL_10_RAW_AVAILABLE"] for row in rows),
        "partial_available": sum(row["PARTIAL_AVAILABLE"] for row in rows),
        "missing_reason_counts": dict(sorted(missing_reasons.items())),
        "supported_industry_missing_reasons": supported_industry_reasons,
        "conditional_industry_5d_missing_reason_counts": {
            reason: conditional_5d_reasons[reason]
            for reason in supported_industry_reasons
        },
        "conditional_industry_20d_missing_reason_counts": {
            reason: conditional_20d_reasons[reason]
            for reason in supported_industry_reasons
        },
        "hsci_history_backfill_decision": {
            "affected_cases": len(affected),
            "affected_years": affected_years,
            "percent_of_438": round(len(affected) / EXPECTED_CASES * 100, 4),
            "development_set_impact": development_affected,
            "validation_set_impact": validation_affected,
            "years_2022_2024_history_fully_covered": all(
                not by_year[year]["missing_because_history_starts_2021_08_19_20d"]
                for year in (2022, 2023, 2024)
            ),
            "temporal_missingness_risk": "HIGH" if affected else "LOW",
            "purchase_recommended": "NO",
            "rationale": (
                "The paid history would repair conditional early-year source coverage, "
                "but cannot make the delivered static industry classifications PIT-safe. "
                "Resolve a listing-time/historically effective classification source first."
            ),
        },
        "pit": "PASS",
        "pit_detail": "HSI/HSCI/HKEX are strict-before; unsafe industry mapping is blocked",
        "future_row_poisoning": "PASS",
        "determinism": "PASS",
        "blind_2025_y_accessed": False,
        "silent_drops": 0,
        "coverage_content_hash": _canonical_hash(rows),
    }


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0]))
        writer.writeheader()
        writer.writerows(rows)


def _write_json(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, ensure_ascii=False, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )


def run(args: argparse.Namespace) -> dict[str, Any]:
    manifest = ExternalMarketSourceManifest.from_path(args.external_manifest)
    market_provider = CompetitionCSVMarketDataProvider(
        args.ipo_data_root, catalog_dir=args.catalog_dir
    )
    metadata = market_provider.iter_listing_metadata()
    if len(metadata) != EXPECTED_CASES or any(item.cohort_year >= 2025 for item in metadata):
        raise ValueError("official Extended cohort must be exactly 438 non-Blind cases")
    case_ids = {item.case_id for item in metadata}
    industry_fields, temporal_audit = load_static_industry_fields(
        args.official_workbook, case_ids
    )
    mapping = load_mapping(args.mapping)
    hsi_provider = CSMARHSIProvider(args.hsi_normalized_csv, args.hsi_manifest)
    hsci_provider = OfficialHSCIProvider(args.hsci_normalized_csv, manifest)
    turnover_provider = OfficialHKEXTurnoverProvider(args.turnover_normalized_csv, manifest)
    prior_ipos, labels = build_labels_and_prior_ipos(metadata, market_provider)

    strict = materialize(
        metadata,
        industry_fields,
        mapping,
        hsi_provider,
        hsci_provider,
        turnover_provider,
        prior_ipos,
        labels,
        provider_cutoffs=True,
    )
    poisoned = materialize(
        metadata,
        industry_fields,
        mapping,
        hsi_provider,
        hsci_provider,
        turnover_provider,
        prior_ipos,
        labels,
        provider_cutoffs=False,
    )
    if strict != poisoned:
        raise AssertionError("future rows changed Extended readiness")
    repeated = materialize(
        metadata,
        industry_fields,
        mapping,
        hsi_provider,
        hsci_provider,
        turnover_provider,
        prior_ipos,
        labels,
        provider_cutoffs=True,
    )
    if strict != repeated:
        raise AssertionError("Extended readiness is not deterministic")
    summary = summarize(strict, manifest, temporal_audit)
    _write_csv(args.runtime_output, strict)
    _write_json(args.summary_output, summary)
    if "INDUSTRY_MAPPING_PIT_BLOCKED" not in args.summary_output.read_text(
        encoding="utf-8"
    ):
        raise AssertionError("UTF-8/content verification failed")
    return summary


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--catalog-dir", type=Path, default=Path("data/catalog"))
    parser.add_argument("--ipo-data-root", type=Path, required=True)
    parser.add_argument("--official-workbook", type=Path, required=True)
    parser.add_argument(
        "--mapping",
        type=Path,
        default=Path("data/catalog/v04_c_hsics_benchmark_mapping_draft.csv"),
    )
    parser.add_argument(
        "--external-manifest",
        type=Path,
        default=Path("data/catalog/v04_c_external_market_source_manifest.json"),
    )
    parser.add_argument("--hsi-normalized-csv", type=Path, required=True)
    parser.add_argument(
        "--hsi-manifest",
        type=Path,
        default=Path("data/catalog/csmar_hsi_source_manifest.json"),
    )
    parser.add_argument("--hsci-normalized-csv", type=Path, required=True)
    parser.add_argument("--turnover-normalized-csv", type=Path, required=True)
    parser.add_argument(
        "--runtime-output",
        type=Path,
        default=Path(
            "data/competition/market_reference/audit/"
            "v04_c_extended_readiness_438.csv"
        ),
    )
    parser.add_argument(
        "--summary-output",
        type=Path,
        default=Path("data/catalog/v04_c_extended_readiness_summary.json"),
    )
    return parser.parse_args()


def main() -> int:
    summary = run(parse_args())
    print(json.dumps(summary, ensure_ascii=False, indent=2, sort_keys=True))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
