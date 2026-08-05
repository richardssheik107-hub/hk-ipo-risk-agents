"""Build the v0.2 competition-data manifest and quality catalogs."""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import sys
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

import fitz
from openpyxl import load_workbook


EXPECTED_YEAR_COUNTS = {
    2020: 138,
    2021: 88,
    2022: 87,
    2023: 63,
    2024: 73,
    2025: 116,
}
EXPECTED_TOTAL = sum(EXPECTED_YEAR_COUNTS.values())
REAL_CASE_CODE = "02410"
OFFICIAL_IPO_WORKBOOK_FILENAME = "HK_Official_Merged_565_First_with_IPO.xlsx"
YEAR_DIRECTORY_PATTERN = re.compile(r"^(20\d{2})_(\d+)份$")
PDF_FILENAME_PATTERN = re.compile(
    r"^(?P<stock_code>\d{5})_(?P<date>\d{2}-\d{2}-\d{4})_"
    r"(?P<company>.+)_(?P<offering>.+)\.pdf$",
    re.IGNORECASE,
)

MANIFEST_FIELDS = (
    "case_id",
    "source_year",
    "source_filename",
    "relative_path",
    "stock_code_raw",
    "stock_code_wind",
    "code_match_method",
    "disclosure_date",
    "company_short_name",
    "offering_type",
    "file_size_bytes",
    "sha256",
    "pdf_page_count",
    "is_text_pdf",
    "parser_status",
    "parser_error_count",
    "eod_available",
    "dataset_split",
    "annotation_status",
    "notes",
)
EOD_COVERAGE_FIELDS = (
    "case_id",
    "source_year",
    "stock_code_raw",
    "stock_code_wind",
    "eod_available",
    "eod_row_count",
    "first_trade_date",
    "last_trade_date",
    "notes",
)
DATASET_SPLIT_FIELDS = (
    "case_id",
    "source_year",
    "stock_code_raw",
    "stock_code_wind",
    "dataset_split",
    "split_reason",
    "is_blind_test",
)
QUALITY_ISSUE_FIELDS = (
    "issue_id",
    "scope",
    "case_id",
    "source_file",
    "issue_code",
    "severity",
    "status",
    "description",
    "recommended_action",
)
OFFICIAL_IPO_BRIDGE_FIELDS = (
    "case_id",
    "source_year",
    "stock_code_raw",
    "stock_code_wind",
    "dataset_split",
    "official_match_status",
    "official_match_method",
    "official_listed_date",
    "institution_id",
    "security_id",
    "selected_name",
    "has_ipo_information",
    "has_institution_info",
    "has_delisted_info",
    "official_ipo_price",
    "official_offer_price",
    "official_delisted_date",
    "official_listing_board_id",
    "official_list_method",
    "official_industry_name",
    "official_funds_raised",
    "official_net_proceed",
    "eod_available",
    "first_eod_trade_date",
    "listed_date_eod_relation",
    "source_workbook",
    "source_workbook_sha256",
    "notes",
)


@dataclass(frozen=True)
class ProspectusSource:
    """One immutable prospectus discovered under a yearly directory."""

    case_id: str
    source_year: int
    source_filename: str
    relative_path: str
    stock_code_raw: str
    stock_code_wind: str
    code_match_method: str
    disclosure_date: str
    company_short_name: str
    offering_type: str
    path: Path


@dataclass
class EODStats:
    """Coverage summary for one Wind stock code."""

    row_count: int = 0
    first_trade_date: str = ""
    last_trade_date: str = ""

    def add(self, trade_date: str) -> None:
        self.row_count += 1
        if trade_date and (not self.first_trade_date or trade_date < self.first_trade_date):
            self.first_trade_date = trade_date
        if trade_date and (not self.last_trade_date or trade_date > self.last_trade_date):
            self.last_trade_date = trade_date


def normalize_stock_code(raw_code: str) -> tuple[str, str]:
    """Normalize a verified five-digit HK equity filename code."""
    if not re.fullmatch(r"\d{5}", raw_code):
        return "", "unmatched_invalid_filename_code"
    if raw_code == "00000":
        return "", "unmatched_invalid_filename_code"
    # Competition filenames use five digits, while Wind HK equity codes use
    # four digits. Remove exactly the filename padding digit; do not strip all
    # leading zeroes (00084 must become 0084.HK, not 84.HK).
    wind_digits = raw_code[1:] if raw_code.startswith("0") else raw_code
    return f"{wind_digits}.HK", "filename_5digit_to_wind_hk_equity"


def dataset_split_for(source_year: int, stock_code_raw: str) -> tuple[str, str]:
    """Apply the frozen chronological split, including the 2410.HK exception."""
    if source_year == 2024 and stock_code_raw == REAL_CASE_CODE:
        return "development_exception", "used_by_v0.2_real_case_001"
    if 2020 <= source_year <= 2023:
        return "development", "chronological_2020_2023"
    if source_year == 2024:
        return "validation", "chronological_2024"
    if source_year == 2025:
        return "blind_test", "chronological_2025_holdout"
    return "quarantined", "outside_frozen_year_range"


def parse_prospectus_path(data_root: Path, path: Path, source_year: int) -> ProspectusSource:
    """Parse stable metadata from one competition prospectus filename."""
    match = PDF_FILENAME_PATTERN.fullmatch(path.name)
    if match is None:
        raise ValueError(f"unexpected prospectus filename: {path.name}")
    stock_code_raw = match.group("stock_code")
    stock_code_wind, method = normalize_stock_code(stock_code_raw)
    disclosure_date = datetime.strptime(match.group("date"), "%d-%m-%Y").date().isoformat()
    return ProspectusSource(
        case_id=f"ipo_{source_year}_{stock_code_raw}",
        source_year=source_year,
        source_filename=path.name,
        relative_path=path.relative_to(data_root).as_posix(),
        stock_code_raw=stock_code_raw,
        stock_code_wind=stock_code_wind,
        code_match_method=method,
        disclosure_date=disclosure_date,
        company_short_name=match.group("company"),
        offering_type=match.group("offering"),
        path=path,
    )


def discover_prospectuses(data_root: Path) -> list[ProspectusSource]:
    """Discover all yearly prospectus directories without touching the blind-set contents."""
    sources: list[ProspectusSource] = []
    for directory in sorted(item for item in data_root.iterdir() if item.is_dir()):
        match = YEAR_DIRECTORY_PATTERN.fullmatch(directory.name)
        if match is None:
            continue
        source_year = int(match.group(1))
        sources.extend(
            parse_prospectus_path(data_root, path, source_year)
            for path in sorted(directory.glob("*.pdf"))
        )
    case_ids = [source.case_id for source in sources]
    if len(case_ids) != len(set(case_ids)):
        raise ValueError("duplicate case_id values discovered")
    return sources


def file_sha256(path: Path, chunk_size: int = 4 * 1024 * 1024) -> str:
    """Return a streaming SHA-256 without loading a large PDF into memory."""
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while chunk := handle.read(chunk_size):
            digest.update(chunk)
    return digest.hexdigest()


def inspect_pdf(path: Path) -> tuple[int, bool, str, int, str]:
    """Validate PDF metadata and sample text without running the production Parser."""
    try:
        with fitz.open(path) as document:
            page_count = document.page_count
            if page_count <= 0:
                return 0, False, "pdf_metadata_failed", 1, "PDF has no pages"
            indexes = sorted({0, page_count // 4, page_count // 2, (page_count * 3) // 4, page_count - 1})
            sampled_characters = sum(len(document.load_page(index).get_text("text").strip()) for index in indexes)
        return page_count, sampled_characters > 0, "not_run", 0, ""
    except Exception as exc:
        return 0, False, "pdf_metadata_failed", 1, f"{type(exc).__name__}: {exc}"


def load_annotation_statuses(catalog_dir: Path) -> dict[str, str]:
    """Reuse existing shadow-review state without making it a B1 dependency."""
    statuses: dict[str, str] = {}
    selection_path = catalog_dir / "shadow_sample_24.csv"
    if selection_path.is_file():
        with selection_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                status = "shadow_manual_reviewed" if row.get("manual_review", "").lower() == "true" else "shadow_automatic"
                statuses[row["stock_code_raw"]] = status
    statuses[REAL_CASE_CODE] = "provisional_gold"
    return statuses


def read_eod_stats(path: Path, target_codes: set[str]) -> tuple[dict[str, EODStats], dict[str, int]]:
    """Stream the 1GB EOD file once and retain only target-code coverage statistics."""
    stats = {code: EODStats() for code in target_codes if code}
    unique_codes: set[str] = set()
    total_rows = 0
    malformed_rows = 0
    with path.open("r", encoding="gb18030", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        expected_columns = len(header)
        try:
            code_index = header.index("S_INFO_WINDCODE")
            date_index = header.index("TRADE_DT")
        except ValueError as exc:
            raise ValueError("EOD file lacks S_INFO_WINDCODE or TRADE_DT") from exc
        for row in reader:
            total_rows += 1
            if len(row) != expected_columns:
                malformed_rows += 1
                continue
            code = row[code_index].strip()
            unique_codes.add(code)
            if code in stats:
                stats[code].add(row[date_index].strip())
    return stats, {
        "rows": total_rows,
        "columns": expected_columns,
        "unique_codes": len(unique_codes),
        "malformed_rows": malformed_rows,
    }


def inspect_csv_shape(path: Path) -> dict[str, int]:
    """Count records and malformed rows using the supplied GB18030 encoding."""
    rows = 0
    malformed_rows = 0
    with path.open("r", encoding="gb18030", newline="") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        expected_columns = len(header)
        for row in reader:
            rows += 1
            if len(row) != expected_columns:
                malformed_rows += 1
    return {"rows": rows, "columns": expected_columns, "malformed_rows": malformed_rows}


def _cell_text(value: object) -> str:
    """Convert an Excel cell to a stable, CSV-friendly representation."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    return str(value).strip()


def _read_workbook_sheet(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    """Read one workbook sheet using its first row as the field header."""
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Workbook is missing required sheet: {sheet_name}")
        worksheet = workbook[sheet_name]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header = [_cell_text(value) for value in next(rows)]
        except StopIteration as exc:
            raise ValueError(f"Workbook sheet is empty: {sheet_name}") from exc
        if not any(header):
            raise ValueError(f"Workbook sheet has no header: {sheet_name}")
        result: list[dict[str, str]] = []
        for values in rows:
            row = {
                column: _cell_text(values[index]) if index < len(values) else ""
                for index, column in enumerate(header)
                if column
            }
            if any(row.values()):
                result.append(row)
        return result
    finally:
        workbook.close()


def _require_workbook_fields(rows: list[dict[str, str]], sheet_name: str, fields: set[str]) -> None:
    if not rows:
        raise ValueError(f"Workbook sheet has no data rows: {sheet_name}")
    missing = fields - set(rows[0])
    if missing:
        raise ValueError(f"Workbook sheet {sheet_name} is missing fields: {sorted(missing)}")


def _listed_date_eod_relation(match_status: str, listed_date: str, first_eod_trade_date: str) -> str:
    if match_status != "matched":
        return "not_applicable_unmatched"
    if not listed_date:
        return "not_comparable_missing_listed_date"
    if not first_eod_trade_date:
        return "no_eod_coverage"
    comparable_eod_date = first_eod_trade_date
    if re.fullmatch(r"\d{8}", comparable_eod_date):
        comparable_eod_date = f"{comparable_eod_date[:4]}-{comparable_eod_date[4:6]}-{comparable_eod_date[6:]}"
    if comparable_eod_date == listed_date:
        return "eod_starts_on_listed_date"
    if comparable_eod_date > listed_date:
        return "eod_starts_after_listed_date"
    return "eod_starts_before_listed_date"


def load_official_ipo_bridge(
    data_root: Path,
    prospectuses: list[ProspectusSource],
    eod_stats: dict[str, EODStats],
) -> list[dict[str, str]]:
    """Build the PDF-to-official-master bridge from the supplied workbook."""
    workbook_path = data_root / OFFICIAL_IPO_WORKBOOK_FILENAME
    if not workbook_path.exists():
        raise FileNotFoundError(f"Missing required official IPO workbook: {workbook_path}")
    match_rows = _read_workbook_sheet(workbook_path, "IPO_565_Match")
    master_rows = _read_workbook_sheet(workbook_path, "Merged_Official_Data")
    _require_workbook_fields(match_rows, "IPO_565_Match", {
        "CaseID", "MatchStatus", "MatchedSymbol", "MatchedListedDate", "MatchedInstitutionID",
        "MatchedSecurityID", "SelectedName", "HasIPOInformation", "HasInstitutionInfo",
        "HasDelistedInfo", "MatchMethod",
    })
    _require_workbook_fields(master_rows, "Merged_Official_Data", {
        "Symbol", "ListedDate", "InstitutionID", "SecurityID", "IPOPrice", "OfferPrice",
        "DelistedDate", "ListingBoardID", "ListMethod", "IndustryName2", "INDUSTRYNAME",
        "FundsRaised", "NetProceed",
    })
    matches_by_case = {row["CaseID"]: row for row in match_rows if row.get("CaseID")}
    source_case_ids = {source.case_id for source in prospectuses}
    if source_case_ids != set(matches_by_case):
        raise ValueError(
            "IPO_565_Match case IDs do not match discovered prospectuses: "
            f"missing={sorted(source_case_ids - set(matches_by_case))}, "
            f"unexpected={sorted(set(matches_by_case) - source_case_ids)}"
        )
    master_by_key: dict[tuple[str, str, str, str], dict[str, str]] = {}
    for row in master_rows:
        # Rows two and three in the supplied workbook describe fields/units.
        if row.get("Symbol") in {"", "股票代码"}:
            continue
        key = (row["Symbol"], row["ListedDate"], row["InstitutionID"], row["SecurityID"])
        if key in master_by_key:
            raise ValueError(f"Duplicate official master key: {key}")
        master_by_key[key] = row
    workbook_hash = file_sha256(workbook_path)
    bridge_rows: list[dict[str, str]] = []
    for source in prospectuses:
        match = matches_by_case[source.case_id]
        status = match["MatchStatus"]
        matched = status == "matched"
        master: dict[str, str] = {}
        if matched:
            key = (match["MatchedSymbol"], match["MatchedListedDate"], match["MatchedInstitutionID"], match["MatchedSecurityID"])
            if key not in master_by_key:
                raise ValueError(f"No official master row for matched IPO case {source.case_id}: {key}")
            master = master_by_key[key]
        eod = eod_stats.get(source.stock_code_wind, EODStats())
        first_eod_trade_date = eod.first_trade_date if eod.row_count else ""
        listed_date = master.get("ListedDate", "") if matched else ""
        bridge_rows.append({
            "case_id": source.case_id, "source_year": str(source.source_year),
            "stock_code_raw": source.stock_code_raw, "stock_code_wind": source.stock_code_wind,
            "dataset_split": dataset_split_for(source.source_year, source.stock_code_raw)[0],
            "official_match_status": status, "official_match_method": match["MatchMethod"],
            "official_listed_date": listed_date,
            "institution_id": match["MatchedInstitutionID"] if matched else "",
            "security_id": match["MatchedSecurityID"] if matched else "",
            "selected_name": match["SelectedName"] if matched else "",
            "has_ipo_information": _source_bool(match["HasIPOInformation"]) if matched else "",
            "has_institution_info": _source_bool(match["HasInstitutionInfo"]) if matched else "",
            "has_delisted_info": _source_bool(match["HasDelistedInfo"]) if matched else "",
            "official_ipo_price": master.get("IPOPrice", ""), "official_offer_price": master.get("OfferPrice", ""),
            "official_delisted_date": master.get("DelistedDate", ""),
            "official_listing_board_id": master.get("ListingBoardID", ""),
            "official_list_method": master.get("ListMethod", ""),
            "official_industry_name": master.get("IndustryName2", "") or master.get("INDUSTRYNAME", ""),
            "official_funds_raised": master.get("FundsRaised", ""), "official_net_proceed": master.get("NetProceed", ""),
            "eod_available": _bool(eod.row_count > 0), "first_eod_trade_date": first_eod_trade_date,
            "listed_date_eod_relation": _listed_date_eod_relation(status, listed_date, first_eod_trade_date),
            "source_workbook": workbook_path.name, "source_workbook_sha256": workbook_hash,
            "notes": "official master match unavailable" if not matched else "",
        })
    return bridge_rows


def _bool(value: bool) -> str:
    return "true" if value else "false"


def _source_bool(value: str) -> str:
    """Normalize boolean-like workbook values for stable catalog output."""
    return "true" if value.strip().lower() in {"true", "1", "yes", "y"} else "false"


def build_rows(
    sources: list[ProspectusSource],
    eod_stats: dict[str, EODStats],
    annotation_statuses: dict[str, str],
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    """Build deterministic manifest, coverage, split and case-level issue rows."""
    manifest_rows: list[dict[str, object]] = []
    coverage_rows: list[dict[str, object]] = []
    split_rows: list[dict[str, object]] = []
    issues: list[dict[str, object]] = []
    for source in sources:
        split, split_reason = dataset_split_for(source.source_year, source.stock_code_raw)
        eod = eod_stats.get(source.stock_code_wind, EODStats())
        eod_available = eod.row_count > 0
        page_count, is_text_pdf, parser_status, parser_errors, pdf_error = inspect_pdf(source.path)
        notes = ["B1 metadata inspection only; production Parser not run"]
        if not eod_available:
            notes.append("no EOD coverage; document-chain/degradation use only")
        if pdf_error:
            notes.append(pdf_error)
        manifest_rows.append(
            {
                "case_id": source.case_id,
                "source_year": source.source_year,
                "source_filename": source.source_filename,
                "relative_path": source.relative_path,
                "stock_code_raw": source.stock_code_raw,
                "stock_code_wind": source.stock_code_wind,
                "code_match_method": source.code_match_method,
                "disclosure_date": source.disclosure_date,
                "company_short_name": source.company_short_name,
                "offering_type": source.offering_type,
                "file_size_bytes": source.path.stat().st_size,
                "sha256": file_sha256(source.path),
                "pdf_page_count": page_count,
                "is_text_pdf": _bool(is_text_pdf),
                "parser_status": parser_status,
                "parser_error_count": parser_errors,
                "eod_available": _bool(eod_available),
                "dataset_split": split,
                "annotation_status": annotation_statuses.get(source.stock_code_raw, "unreviewed"),
                "notes": "; ".join(notes),
            }
        )
        coverage_rows.append(
            {
                "case_id": source.case_id,
                "source_year": source.source_year,
                "stock_code_raw": source.stock_code_raw,
                "stock_code_wind": source.stock_code_wind,
                "eod_available": _bool(eod_available),
                "eod_row_count": eod.row_count,
                "first_trade_date": eod.first_trade_date,
                "last_trade_date": eod.last_trade_date,
                "notes": "" if eod_available else "document-chain/degradation use only",
            }
        )
        split_rows.append(
            {
                "case_id": source.case_id,
                "source_year": source.source_year,
                "stock_code_raw": source.stock_code_raw,
                "stock_code_wind": source.stock_code_wind,
                "dataset_split": split,
                "split_reason": split_reason,
                "is_blind_test": _bool(split == "blind_test"),
            }
        )
        if not eod_available:
            issues.append(
                {
                    "scope": "case",
                    "case_id": source.case_id,
                    "source_file": "hkshareeodprices.csv",
                    "issue_code": "EOD_NOT_AVAILABLE",
                    "severity": "warning",
                    "status": "accepted_degradation",
                    "description": f"{source.stock_code_wind} has no matching EOD rows",
                    "recommended_action": "Use for document-chain and degradation tests only",
                }
            )
        if parser_status == "pdf_metadata_failed":
            issues.append(
                {
                    "scope": "case",
                    "case_id": source.case_id,
                    "source_file": source.relative_path,
                    "issue_code": "PDF_METADATA_FAILED",
                    "severity": "error",
                    "status": "open",
                    "description": pdf_error,
                    "recommended_action": "Inspect or replace the source PDF",
                }
            )
    return manifest_rows, coverage_rows, split_rows, issues


def dataset_issues(data_root: Path, shapes: dict[str, dict[str, int]]) -> list[dict[str, object]]:
    """Record known dataset-level limitations without guessing missing values."""
    security = shapes["hksharedescription.csv"]
    security_path = data_root / "hksharedescription.csv"
    truncated = security_path.stat().st_size == 262_144 or security["malformed_rows"] > 0
    issues: list[dict[str, object]] = []
    if truncated:
        issues.append(
            {
                "scope": "dataset",
                "case_id": "",
                "source_file": "hksharedescription.csv",
                "issue_code": "SECURITY_MASTER_TRUNCATED",
                "severity": "critical",
                "status": "quarantined",
                "description": (
                    f"Security master is {security_path.stat().st_size} bytes with "
                    f"{security['rows']} records and {security['malformed_rows']} malformed rows; truncation is suspected"
                ),
                "recommended_action": "Do not derive listing date, issue price or company mapping until replaced",
            }
        )
    issues.append(
        {
            "scope": "dataset",
            "case_id": "",
            "source_file": "hkshareeodprices.csv",
            "issue_code": "EOD_AMOUNT_UNIT_UNCONFIRMED",
            "severity": "warning",
            "status": "open",
            "description": "The unit of S_DQ_AMOUNT is not confirmed for feature engineering",
            "recommended_action": "Confirm against the official data dictionary before v0.4",
        }
    )
    return issues


def official_bridge_issues(bridge_rows: list[dict[str, str]]) -> list[dict[str, object]]:
    """Record explicit gaps in the supplied official IPO master bridge."""
    issues: list[dict[str, object]] = []
    for row in bridge_rows:
        if row["official_match_status"] != "matched":
            issues.append({
                "scope": "case", "case_id": row["case_id"],
                "source_file": OFFICIAL_IPO_WORKBOOK_FILENAME,
                "issue_code": "OFFICIAL_IPO_MASTER_MATCH_MISSING", "severity": "warning",
                "status": "open", "description": "No official IPO master row was supplied for this prospectus case",
                "recommended_action": "Retain as unavailable; review the official-source matching workflow before enrichment",
            })
        elif row["has_institution_info"].lower() != "true":
            issues.append({
                "scope": "case", "case_id": row["case_id"],
                "source_file": OFFICIAL_IPO_WORKBOOK_FILENAME,
                "issue_code": "OFFICIAL_IPO_INSTITUTION_INFO_MISSING", "severity": "warning",
                "status": "accepted_degradation", "description": "Official IPO record is matched but InstitutionInfo is unavailable",
                "recommended_action": "Use available IPO fields only; do not infer missing company-profile values",
            })
    return issues


def write_csv(path: Path, rows: Iterable[dict[str, object]], fields: tuple[str, ...]) -> None:
    """Write deterministic Excel-friendly UTF-8 CSV output."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows({field: row.get(field, "") for field in fields} for row in rows)


def write_docs(
    docs_dir: Path,
    manifest_rows: list[dict[str, object]],
    issues: list[dict[str, object]],
    shapes: dict[str, dict[str, int]],
    official_bridge_rows: list[dict[str, str]],
) -> None:
    """Generate concise source-backed overview and quality documentation."""
    year_counts = Counter(int(row["source_year"]) for row in manifest_rows)
    split_counts = Counter(str(row["dataset_split"]) for row in manifest_rows)
    eod_available = sum(row["eod_available"] == "true" for row in manifest_rows)
    text_pdfs = sum(row["is_text_pdf"] == "true" for row in manifest_rows)
    overview = f"""# 赛事数据概览

> 本文件由 `scripts/build_competition_manifest.py` 生成。原始数据只读使用，不进入Git。

## 数据规模

| 数据 | 规模 | 编码/形式 |
|---|---:|---|
| 招股书 | {len(manifest_rows)}份 | 按年份目录存放的PDF |
| 公司资料 | {shapes['hkcompanyinfo.csv']['rows']}行，{shapes['hkcompanyinfo.csv']['columns']}列 | GB18030 CSV |
| 证券资料 | {shapes['hksharedescription.csv']['rows']}行，{shapes['hksharedescription.csv']['columns']}列 | GB18030 CSV；隔离使用 |
| 日行情 | {shapes['hkshareeodprices.csv']['rows']}行，{shapes['hkshareeodprices.csv']['columns']}列 | GB18030 CSV |
| 行情代码 | {shapes['hkshareeodprices.csv']['unique_codes']}个 | `S_INFO_WINDCODE` |

## 招股书年度分布

| 年份 | 数量 | 数据用途 |
|---:|---:|---|
"""
    for year in sorted(year_counts):
        split_label = "开发" if year <= 2023 else "验证" if year == 2024 else "盲测"
        overview += f"| {year} | {year_counts[year]} | {split_label} |\n"
    overview += f"""

## 固定数据集划分

- 开发集（2020—2023）：{split_counts['development']}份。
- 2024验证集：{split_counts['validation']}份。
- 开发例外：{split_counts['development_exception']}份，即2410.HK，已用于v0.2真实案例。
- 2025盲测集：{split_counts['blind_test']}份，禁止用于调规则、选特征或调参。

## 覆盖与可用性

- 有日行情覆盖：{eod_available}/{len(manifest_rows)}份。
- 无日行情覆盖：{len(manifest_rows) - eod_available}份，仅用于文档链路和降级测试。
- 抽样可识别文本的PDF：{text_pdfs}/{len(manifest_rows)}份。
- B1仅完成PDF元数据和抽样文本检查；`parser_status=not_run`表示未批量运行生产Parser。

## 关联规则

- 招股书股票代码来自受控文件名，原始五位代码保存在`stock_code_raw`。
- 只有通过五位数字校验后才生成`stock_code_wind`，匹配失败不会伪装为成功。
- 证券主表疑似截断，已隔离；当前不使用它生成上市日期、发行价或公司映射。
- `disclosure_date`来自招股书文件名，不得当作上市日期。
"""
    quality = """# 赛事数据质量报告

> 本文件由 `scripts/build_competition_manifest.py` 生成。详细逐项记录见
> `data/catalog/data_quality_issues.csv`。

## 结论

赛事数据可以支持v0.2文档链路和影子测试，但尚不能直接支持正式上市后风险标签或市场模型。
证券主表在补齐前必须隔离；成交金额单位在确认前不得进入特征工程；无行情样本只能用于降级测试。

## 问题汇总

| 问题代码 | 数量 | 最高严重度 | 当前处理 |
|---|---:|---|---|
"""
    counts = Counter(str(issue["issue_code"]) for issue in issues)
    severity_rank = {"info": 0, "warning": 1, "error": 2, "critical": 3}
    for issue_code in sorted(counts):
        matching = [issue for issue in issues if issue["issue_code"] == issue_code]
        severity = max((str(issue["severity"]) for issue in matching), key=lambda value: severity_rank[value])
        statuses = ", ".join(sorted({str(issue["status"]) for issue in matching}))
        quality += f"| {issue_code} | {counts[issue_code]} | {severity} | {statuses} |\n"
    quality += """

## 使用限制

1. 2025年116份招股书是盲测集，不得用于调试Retriever、构造规则或模型调参。
2. `hksharedescription.csv`处于`quarantined`状态，不得据此生成上市日期和发行价。
3. 招股书披露日期不是上市日期。
4. `S_DQ_AMOUNT`单位确认前，不构造成交金额类特征。
5. 无行情覆盖的样本不构造价格标签，只用于文档链路及降级测试。
6. 所有未匹配、坏行和缺失值必须保留问题记录，不得人工猜测后回填。
"""
    official_matched = sum(row["official_match_status"] == "matched" for row in official_bridge_rows)
    official_unmatched = len(official_bridge_rows) - official_matched
    eod_before_listed = sum(
        row["listed_date_eod_relation"] == "eod_starts_before_listed_date"
        for row in official_bridge_rows
    )
    overview += f"""

## IPO官方主数据桥接

- `HK_Official_Merged_565_First_with_IPO.xlsx` 已作为只读的原始输入；桥接目录为 `data/catalog/ipo_official_master_bridge.csv`。
- 招股书案例与官方主数据：{official_matched}/{len(official_bridge_rows)} 个 `matched`，另有 {official_unmatched} 个 `manifest_only_placeholder`，不得补猜公司、上市日期或发行信息。
- `official_listed_date` 是来源工作簿提供的上市日期；必须结合 `first_eod_trade_date` 和 `listed_date_eod_relation` 使用，不能把日行情最早日期自动当作上市日期。
- 有 {eod_before_listed} 个已匹配案例的日行情早于工作簿上市日期；在进入建模标签或时间窗前必须人工复核该日期关系。
"""
    docs_dir.mkdir(parents=True, exist_ok=True)
    (docs_dir / "COMPETITION_DATA_OVERVIEW.md").write_text(overview, encoding="utf-8", newline="\n")
    (docs_dir / "DATA_QUALITY_REPORT.md").write_text(quality, encoding="utf-8", newline="\n")


def build(data_root: Path, catalog_dir: Path, docs_dir: Path) -> dict[str, int]:
    """Build all B1 deliverables from a read-only competition data root."""
    required_files = (
        "hkcompanyinfo.csv", "hksharedescription.csv", "hkshareeodprices.csv",
        OFFICIAL_IPO_WORKBOOK_FILENAME,
    )
    missing = [name for name in required_files if not (data_root / name).is_file()]
    if missing:
        raise FileNotFoundError(f"missing competition data files: {', '.join(missing)}")

    sources = discover_prospectuses(data_root)
    target_codes = {source.stock_code_wind for source in sources if source.stock_code_wind}
    eod_stats, eod_shape = read_eod_stats(data_root / "hkshareeodprices.csv", target_codes)
    shapes = {
        "hkcompanyinfo.csv": inspect_csv_shape(data_root / "hkcompanyinfo.csv"),
        "hksharedescription.csv": inspect_csv_shape(data_root / "hksharedescription.csv"),
        "hkshareeodprices.csv": eod_shape,
    }
    annotations = load_annotation_statuses(catalog_dir)
    manifest_rows, coverage_rows, split_rows, issues = build_rows(sources, eod_stats, annotations)
    official_bridge_rows = load_official_ipo_bridge(data_root, sources, eod_stats)
    issues.extend(dataset_issues(data_root, shapes))
    issues.extend(official_bridge_issues(official_bridge_rows))
    for index, issue in enumerate(issues, start=1):
        issue["issue_id"] = f"DQ-{index:04d}"

    write_csv(catalog_dir / "ipo_prospectus_manifest.csv", manifest_rows, MANIFEST_FIELDS)
    write_csv(catalog_dir / "eod_coverage_report.csv", coverage_rows, EOD_COVERAGE_FIELDS)
    write_csv(catalog_dir / "dataset_split.csv", split_rows, DATASET_SPLIT_FIELDS)
    write_csv(catalog_dir / "ipo_official_master_bridge.csv", official_bridge_rows, OFFICIAL_IPO_BRIDGE_FIELDS)
    write_csv(catalog_dir / "data_quality_issues.csv", issues, QUALITY_ISSUE_FIELDS)
    write_docs(docs_dir, manifest_rows, issues, shapes, official_bridge_rows)
    return {
        "prospectuses": len(manifest_rows),
        "eod_available": sum(row["eod_available"] == "true" for row in manifest_rows),
        "eod_missing": sum(row["eod_available"] == "false" for row in manifest_rows),
        "official_matched": sum(row["official_match_status"] == "matched" for row in official_bridge_rows),
        "official_unmatched": sum(row["official_match_status"] != "matched" for row in official_bridge_rows),
        "quality_issues": len(issues),
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--data-root",
        type=Path,
        default=Path(os.environ.get("IPO_RISK_COMPETITION_DATA_ROOT", "data/competition")),
    )
    parser.add_argument("--catalog-dir", type=Path, default=Path("data/catalog"))
    parser.add_argument("--docs-dir", type=Path, default=Path("docs"))
    args = parser.parse_args()
    if not args.data_root.is_dir():
        parser.error(f"competition data root not found: {args.data_root}")
    summary = build(args.data_root, args.catalog_dir, args.docs_dir)
    print(" ".join(f"{key}={value}" for key, value in summary.items()))
    return 0


if __name__ == "__main__":
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    raise SystemExit(main())
